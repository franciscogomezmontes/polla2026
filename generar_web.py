#!/usr/bin/env python3
"""
generar_web.py - La Mejor Polla Mundialista 2026
=================================================
Lee el admin Excel (hoja Posiciones, bloque ampliado) y genera index.html
con los datos incrustados. Sube index.html a GitHub y el ranking se
actualiza en el sitio.

Esta version incluye automaticamente:
  - Links en cada nombre hacia su pagina de predicciones
    (ranking-animado/participante.html?id=<slug>)
  - Boton hacia la grafica animada (ranking-animado/grafica.html)

Los slugs se leen de ranking-animado/data/predicciones/_indice.json,
que genera exportar_predicciones.py. Si ese archivo no existe o un
participante no tiene slug todavia, el link simplemente no se genera
para ese caso (sin romper nada).

Uso:
    python generar_web.py
    python generar_web.py --admin "Admin_POLLA_FIFA_WORLD_CUP_2026_r00.xlsx"
    python generar_web.py --jornada "Jornada 3 - 14 Jun 2026"
    python generar_web.py --output "index.html"
"""
import argparse, json, os, sys
from datetime import datetime
import openpyxl
from openpyxl.cell.cell import MergedCell

ADMIN_DEFAULT = "Admin_POLLA_FIFA_WORLD_CUP_2026_r00.xlsx"
INDICE_PREDICCIONES_DEFAULT = os.path.join(
    "ranking-animado", "data", "predicciones", "_indice.json"
)

# Hoja Posiciones - bloque ampliado (columnas base 1)
COL_RANKING    = 8    # Ranking Polla (actual)
COL_RANK_ANT   = 9    # Ranking Anterior
COL_MOV        = 10   # Mov. Rank.
COL_PUNTOS     = 12   # Puntos Totales
COL_PUNTOS_ANT = 13   # Puntos Anterior Jornada
COL_MOV_PUNTOS = 14   # Mov. Puntos (diferencia)
COL_NOMBRE     = 15   # Nombre
FILA_INICIO    = 4

def leer_ranking(admin_path):
    wb = openpyxl.load_workbook(admin_path, data_only=True)
    ws = wb['Posiciones']

    participantes = []
    for row in range(FILA_INICIO, ws.max_row + 1):
        c_nom = ws.cell(row, COL_NOMBRE)
        if isinstance(c_nom, MergedCell) or not c_nom.value:
            continue

        ranking  = ws.cell(row, COL_RANKING).value
        rank_ant = ws.cell(row, COL_RANK_ANT).value
        mov_raw  = ws.cell(row, COL_MOV).value
        puntos   = ws.cell(row, COL_PUNTOS).value
        pts_ant  = ws.cell(row, COL_PUNTOS_ANT).value
        mov_pts  = ws.cell(row, COL_MOV_PUNTOS).value

        if mov_raw is not None:
            mov = int(mov_raw)
        elif ranking is not None and rank_ant is not None:
            mov = int(rank_ant) - int(ranking)
        else:
            mov = 0

        participantes.append({
            "ranking":    int(ranking) if ranking is not None else 0,
            "nombre":     str(c_nom.value).strip(),
            "puntos":     int(puntos) if puntos is not None else 0,
            "puntos_ant": int(pts_ant) if pts_ant is not None else 0,
            "diff_pts":   int(mov_pts) if mov_pts is not None else 0,
            "mov":        mov,
        })

    participantes.sort(key=lambda x: (x["ranking"], -x["puntos"]))
    return participantes


def leer_slugs(ruta_indice):
    """
    Lee ranking-animado/data/predicciones/_indice.json y devuelve un dict
    {nombre: slug}. Si el archivo no existe, devuelve un dict vacio sin
    lanzar error: el index.html se genera igual, simplemente sin links
    de participante (se podran agregar despues de correr el exportador
    de predicciones).
    """
    if not os.path.exists(ruta_indice):
        print(f"Aviso: no se encontro {ruta_indice}")
        print("  Los nombres se generaran sin link a su pagina de predicciones.")
        print("  Corre exportar_predicciones.py para habilitar los links.")
        return {}

    with open(ruta_indice, encoding="utf-8") as f:
        indice = json.load(f)

    return {p["nombre"]: p["slug"] for p in indice.get("participantes", [])}


HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>La Mejor Polla Mundialista 2026</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800;900&family=Barlow:wght@400;500;600&display=swap');
  :root {
    --rojo:#C8102E;--azul:#003F87;--verde:#1A7A4A;--oro:#F5A623;
    --plata:#A0A8B0;--bronce:#CD7F32;--fondo:#0A0E1A;--panel:#111827;
    --borde:#1E2A40;--texto:#F0F4FF;--muted:#6B7A99;
    --sube:#22C55E;--baja:#EF4444;--igual:#6B7A99;
  }
  * { margin:0;padding:0;box-sizing:border-box; }
  body { background:var(--fondo);color:var(--texto);font-family:'Barlow',sans-serif;min-height:100vh; }
  header {
    background:linear-gradient(135deg,var(--azul) 0%,#001F5B 60%,#0A0E1A 100%);
    border-bottom:3px solid var(--rojo);
    padding:20px 24px 16px;text-align:center;position:relative;overflow:hidden;
  }
  header::before {
    content:'';position:absolute;inset:0;pointer-events:none;
    background:repeating-linear-gradient(45deg,transparent,transparent 20px,rgba(255,255,255,.02) 20px,rgba(255,255,255,.02) 21px);
  }
  .h-trophy{font-size:2.5rem;line-height:1;margin-bottom:6px}
  .h-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(1.4rem,5vw,2.2rem);font-weight:900;letter-spacing:.04em;text-transform:uppercase;line-height:1.1}
  .h-sub{font-family:'Barlow Condensed',sans-serif;font-size:clamp(.8rem,3vw,1rem);font-weight:600;color:var(--oro);letter-spacing:.12em;text-transform:uppercase;margin-top:4px}
  .h-jornada{display:inline-block;background:var(--rojo);color:#fff;font-family:'Barlow Condensed',sans-serif;font-weight:700;font-size:.85rem;letter-spacing:.1em;text-transform:uppercase;padding:4px 14px;border-radius:20px;margin-top:10px}
  .h-grafica-link{display:inline-block;margin-top:10px;color:#F0F4FF;font-size:.75rem;text-decoration:none;border:1px solid rgba(255,255,255,.25);padding:4px 12px;border-radius:20px}
  .h-grafica-link:hover{background:rgba(255,255,255,.08)}
  main{max-width:640px;margin:0 auto;padding:16px 12px 40px}
  table{width:100%;border-collapse:collapse}
  thead th{font-family:'Barlow Condensed',sans-serif;font-size:.7rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--muted);padding:8px 6px;text-align:center;border-bottom:1px solid var(--borde)}
  thead th.th-n{text-align:left;padding-left:10px}
  tbody tr{border-bottom:1px solid var(--borde);transition:background .15s}
  tbody tr:hover{background:rgba(255,255,255,.03)}
  tr.p1{background:rgba(245,166,35,.08)}
  tr.p2{background:rgba(160,168,176,.06)}
  tr.p3{background:rgba(205,127,50,.06)}
  td{padding:10px 6px;text-align:center;font-size:.92rem}
  .td-n{text-align:left;padding-left:10px}
  .td-pos{font-family:'Barlow Condensed',sans-serif;font-size:1.1rem;font-weight:800;width:36px;min-width:36px}
  .c1{color:var(--oro)}.c2{color:var(--plata)}.c3{color:var(--bronce)}
  .td-mov{width:54px;min-width:54px}
  .pill{display:inline-flex;align-items:center;gap:3px;font-family:'Barlow Condensed',sans-serif;font-size:.78rem;font-weight:700;padding:2px 7px;border-radius:20px}
  .up{background:rgba(34,197,94,.15);color:var(--sube)}
  .dn{background:rgba(239,68,68,.15);color:var(--baja)}
  .eq{background:rgba(107,122,153,.15);color:var(--igual)}
  .nom{font-weight:600;font-size:.95rem;line-height:1.2}
  .nom-link{color:inherit;text-decoration:none}
  .nom-link:hover{text-decoration:underline}
  .dpts{font-size:.72rem;color:var(--muted);margin-top:1px}
  .dpos{color:var(--sube)}.dneg{color:var(--baja)}
  .td-pts{font-family:'Barlow Condensed',sans-serif;font-size:1.3rem;font-weight:800;width:52px;min-width:52px}
  .sep td{padding:2px 0;background:var(--borde);font-size:0}
  footer{text-align:center;color:var(--muted);font-size:.72rem;padding:24px 12px 12px;letter-spacing:.05em}
</style>
</head>
<body>
<header>
  <div class="h-trophy">&#x1F3C6;</div>
  <div class="h-title">La Mejor Polla Mundialista</div>
  <div class="h-sub">FIFA World Cup 2026 &middot; Canada &middot; M&eacute;xico &middot; USA</div>
  <div class="h-jornada">__JORNADA__</div>
  <div><a class="h-grafica-link" href="ranking-animado/grafica.html">Ver ranking animado &rarr;</a></div>
</header>
<main>
<table>
  <thead><tr>
    <th>#</th><th>Mov.</th><th class="th-n">Participante</th><th>Pts</th>
  </tr></thead>
  <tbody id="tbody"></tbody>
</table>
</main>
<footer id="footer"></footer>
<script>
var DATA = __DATA__;
var SLUGS = __SLUGS__;
var p = DATA.participantes;
var ts = new Date(DATA.generado);
document.getElementById('footer').textContent =
  'Actualizado: ' + ts.toLocaleDateString('es-CO',{day:'2-digit',month:'short',year:'numeric'}) +
  ' ' + ts.toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
function pill(m){
  if(m>0) return '<span class="pill up">&#9650;+'+m+'</span>';
  if(m<0) return '<span class="pill dn">&#9660;'+m+'</span>';
  return '<span class="pill eq">=</span>';
}
function medal(r){return r===1?'&#x1F947; ':r===2?'&#x1F948; ':r===3?'&#x1F949; ':''}
function pc(r){return r===1?' c1':r===2?' c2':r===3?' c3':''}
function rc(r){return r<=3?' p'+r:''}
function dpts(d){
  if(!d) return '';
  var cl=d>0?' dpos':' dneg', s=d>0?'+':'';
  return '<div class="dpts'+cl+'">'+s+d+' pts esta jornada</div>';
}
function nomLink(nombre){
  var slug = SLUGS[nombre];
  if(!slug) return nombre;
  return '<a class="nom-link" href="ranking-animado/participante.html?id='+slug+'">'+nombre+'</a>';
}
var prev = 0, html = '';
p.forEach(function(x){
  if(prev && prev!==x.ranking && (x.ranking>3 || prev<=3 || x.ranking>10 && prev<=10))
    html += '<tr class="sep"><td colspan="4"></td></tr>';
  html += '<tr class="'+rc(x.ranking)+'">'+
    '<td class="td-pos'+pc(x.ranking)+'">'+x.ranking+'</td>'+
    '<td class="td-mov">'+pill(x.mov)+'</td>'+
    '<td class="td-n"><div class="nom">'+medal(x.ranking)+nomLink(x.nombre)+'</div>'+dpts(x.diff_pts)+'</td>'+
    '<td class="td-pts'+pc(x.ranking)+'">'+x.puntos+'</td>'+
  '</tr>';
  prev = x.ranking;
});
document.getElementById('tbody').innerHTML = html;
</script>
</body>
</html>'''

def generar_html(datos, jornada, slugs):
    data_json = json.dumps(datos, ensure_ascii=False)
    slugs_json = json.dumps(slugs, ensure_ascii=False)
    html = HTML_TEMPLATE
    html = html.replace('__JORNADA__', jornada)
    html = html.replace('__DATA__', data_json)
    html = html.replace('__SLUGS__', slugs_json)
    return html

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--admin",   default=ADMIN_DEFAULT)
    parser.add_argument("--jornada", default="")
    parser.add_argument("--output",  default="index.html")
    parser.add_argument("--indice-predicciones", default=INDICE_PREDICCIONES_DEFAULT,
                         help="Ruta a _indice.json generado por exportar_predicciones.py")
    args = parser.parse_args()

    if not os.path.exists(args.admin):
        print(f"No se encontro: {args.admin}"); sys.exit(1)

    print(f"Leyendo {args.admin}...")
    participantes = leer_ranking(args.admin)

    if not participantes:
        print("No se encontraron participantes con datos calculados.")
        print("Abre el admin en Excel, presiona F9, guarda, y vuelve a correr el script.")
        sys.exit(1)

    slugs = leer_slugs(args.indice_predicciones)
    if slugs:
        sin_slug = [p["nombre"] for p in participantes if p["nombre"] not in slugs]
        if sin_slug:
            print(f"Aviso: {len(sin_slug)} participante(s) sin slug encontrado, no tendran link:")
            for n in sin_slug:
                print(f"  - {n}")

    label = args.jornada or f"Actualizado {datetime.now().strftime('%d %b %Y %H:%M')}"
    datos = {
        "jornada":       label,
        "generado":      datetime.now().isoformat(),
        "participantes": participantes,
    }

    html = generar_html(datos, label, slugs)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"OK: {args.output} generado con {len(participantes)} participantes")
    print(f"Jornada: {label}")
    print(f"Links a predicciones: {len(slugs)} slugs cargados")
    print()
    print("Siguiente paso - sube index.html a GitHub:")
    print("  1. Ve a tu repositorio en github.com")
    print("  2. Arrastra index.html al repositorio")
    print("  3. Haz clic en 'Commit changes'")
    print("  El link se actualiza en ~30 segundos")

if __name__ == "__main__":
    main()
