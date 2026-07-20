"""
exportar_predicciones.py

Lee la hoja 'Puntos' del Admin de la Polla Mundialista y genera un archivo
JSON por participante con todas sus predicciones: partidos de grupo,
clasificados de grupo, eliminatoria completa, y premios FIFA.

NO recalcula nada: copia los valores que el Admin ya tiene (predicciones,
resultados reales, puntos obtenidos). El Admin sigue siendo la unica
fuente de verdad.

Uso:
    python exportar_predicciones.py "Admin_POLLA_FIFA_WORLD_CUP_2026_r00.xlsx" "salida/predicciones/"

Genera un archivo <slug>.json por cada participante dentro de la carpeta
de salida indicada, usando el mismo slug que exportar_ranking.py para que
las paginas de ranking y de predicciones queden enlazadas.
"""

import sys
import os
import json
import re
from datetime import datetime, time as dtime
import openpyxl


SHEET = "Puntos"
FILA_NOMBRE_PART = 3
FILA_TIPO_COL = 4
PRIMER_COL_PARTICIPANTE = 11
COLS_POR_BLOQUE = 4

COL_NUM_PARTIDO = 1
COL_GRUPO = 2
COL_FECHA = 3
COL_HORA = 4
COL_EQUIPO1 = 5
COL_VS = 6
COL_EQUIPO2 = 7
COL_REAL_LOCAL = 8
COL_REAL_VISITANTE = 9
COL_REAL_GANADOR = 10

FILA_INICIO_GRUPOS = 5
FILA_FIN_GRUPOS = 112

FASES_ELIMINATORIA = {
    "DIECISEISAVOS": "Dieciseisavos",
    "OCTAVOS": "Octavos",
    "CUARTOS": "Cuartos",
    "SEMIS": "Semifinales",
    "3ro": "3er y 4to puesto",
    "1ro": "Final",
}

PREMIOS_LABELS = {
    "Goleador del Torneo": "Goleador del Torneo (Bota de Oro)",
    "Mejor Jugador": "Mejor Jugador (Balon de Oro)",
    "Mejor Jugador Joven": "Mejor Jugador Joven",
    "Mejor Arquero": "Mejor Arquero (Guante de Oro)",
    "Equipo Fair Play": "Equipo Fair Play",
    "Equipo más goles anotados": "Equipo que anoto mas goles",
    "Equipo más goles recibidos": "Equipo al que le anotaron mas goles",
    "Equipo sorpresa (menor ranking)": "Equipo sorpresa",
}

RIVAL_MEJOR_TERCERO = {
    "1E": "3ABCDF",
    "1I": "3CDFGH",
    "1A": "3CEFHI",
    "1L": "3EHIJK",
    "1D": "3BEFIJ",
    "1G": "3AEHIJ",
    "1B": "3EFGIJ",
    "1K": "3DEIJL",
}


def slugify(nombre):
    s = nombre.strip().lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
           .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def fmt_fecha_hora(fecha, hora):
    out = ""
    if isinstance(fecha, datetime):
        out += fecha.strftime("%d/%m/%Y")
    if isinstance(hora, dtime):
        out += " " + hora.strftime("%H:%M")
    return out.strip() or None


def detectar_participantes(ws):
    """
    Devuelve lista de dicts {nombre, col_inicio} para cada bloque de
    4 columnas a partir de PRIMER_COL_PARTICIPANTE. No asume un numero
    fijo de participantes: recorre hasta que no encuentre mas nombres.
    """
    participantes = []
    c = PRIMER_COL_PARTICIPANTE
    while c <= ws.max_column:
        nombre = ws.cell(row=FILA_NOMBRE_PART, column=c).value
        if not nombre:
            break
        participantes.append({"nombre": str(nombre).strip(), "col_inicio": c})
        c += COLS_POR_BLOQUE
    return participantes


def extraer_partidos_grupo(ws, col_inicio):
    partidos = []
    grupo_actual = None
    r = FILA_INICIO_GRUPOS
    while r <= FILA_FIN_GRUPOS:
        grupo_cell = ws.cell(row=r, column=COL_GRUPO).value
        if grupo_cell:
            grupo_actual = grupo_cell

        num_partido = ws.cell(row=r, column=COL_NUM_PARTIDO).value
        eq1 = ws.cell(row=r, column=COL_EQUIPO1).value

        if num_partido is not None and eq1 is not None:
            eq2 = ws.cell(row=r, column=COL_EQUIPO2).value
            fecha = ws.cell(row=r, column=COL_FECHA).value
            hora = ws.cell(row=r, column=COL_HORA).value
            real_local = ws.cell(row=r, column=COL_REAL_LOCAL).value
            real_visit = ws.cell(row=r, column=COL_REAL_VISITANTE).value

            pred_local = ws.cell(row=r, column=col_inicio).value
            pred_visit = ws.cell(row=r, column=col_inicio + 1).value
            pred_ganador = ws.cell(row=r, column=col_inicio + 2).value
            pred_puntos = ws.cell(row=r, column=col_inicio + 3).value

            jugado = real_local is not None and real_visit is not None

            partidos.append({
                "grupo": grupo_actual,
                "equipo_local": eq1,
                "equipo_visitante": eq2,
                "fecha": fmt_fecha_hora(fecha, hora),
                "jugado": jugado,
                "resultado_real": {"local": real_local, "visitante": real_visit} if jugado else None,
                "prediccion": {"local": pred_local, "visitante": pred_visit, "ganador": pred_ganador},
                "puntos": pred_puntos if jugado else None,
            })
        elif grupo_cell is None:
            texto_clasif = ws.cell(row=r, column=COL_EQUIPO1).value
            if texto_clasif and re.match(r"^\d[A-L]$", str(texto_clasif)):
                pred_ganador = ws.cell(row=r, column=col_inicio + 2).value
                pred_puntos = ws.cell(row=r, column=col_inicio + 3).value
                partidos.append({
                    "tipo": "clasificado_grupo",
                    "posicion": texto_clasif,
                    "prediccion": pred_ganador,
                    "puntos": pred_puntos,
                })
        r += 1
    return partidos


FECHAS_ELIMINATORIA = {
    # Dieciseisavos (partidos 73-88)
    73: '2026-06-28', 74: '2026-06-29', 75: '2026-06-29', 76: '2026-06-29',
    77: '2026-06-30', 78: '2026-06-30', 79: '2026-06-30', 80: '2026-07-01',
    81: '2026-07-01', 82: '2026-07-01', 83: '2026-07-02', 84: '2026-07-02',
    85: '2026-07-02', 86: '2026-07-03', 87: '2026-07-03', 88: '2026-07-03',
    # Octavos (partidos 89-96)
    89: '2026-07-04', 90: '2026-07-04', 91: '2026-07-05', 92: '2026-07-05',
    93: '2026-07-06', 94: '2026-07-06', 95: '2026-07-07', 96: '2026-07-07',
    # Cuartos (partidos 97-100)
    97: '2026-07-09', 98: '2026-07-10', 99: '2026-07-11', 100: '2026-07-11',
    # Semis (partidos 101-102)
    101: '2026-07-14', 102: '2026-07-15',
}


def construir_mapa_clasificados(ws, partidos_pdf):
    """
    Construye {fila_partido_admin: fila_clasificado_admin} para todos los
    partidos de eliminatoria que tienen clasificado asociado.

    Estrategia unificada: cruce por (fecha_colombia, hora_ET) como clave
    compuesta. Funciona sin importar si el Admin tiene codigos (2A, 1C) o
    nombres reales de equipo (Brazil, Germany).
    """
    # Mapa numero_oficial -> fila_clasificado leyendo etiquetas del Admin
    clasif_por_num = {}
    for r in range(129, 175):
        label = ws.cell(row=r, column=COL_EQUIPO1).value
        if not label or not isinstance(label, str):
            continue
        label = label.strip()
        if label.startswith('Clasificado M'):
            try:
                clasif_por_num[int(label.replace('Clasificado M', ''))] = r
            except ValueError:
                pass
        elif label.startswith('Clasificado W'):
            try:
                clasif_por_num[int(label.replace('Clasificado W', ''))] = r
            except ValueError:
                pass
        elif label == 'W101':
            clasif_por_num[101] = r
        elif label == 'W102':
            clasif_por_num[102] = r

    # Mapa (fecha_str, hora_ET) -> numero_oficial para todos los partidos de eliminatoria
    mapa_fecha_hora_a_num = {}
    for p in partidos_pdf:
        num = p['numero']
        if num not in FECHAS_ELIMINATORIA:
            continue
        fecha_str = FECHAS_ELIMINATORIA[num]
        hora_et = p['hora']
        mapa_fecha_hora_a_num[(fecha_str, hora_et)] = num

    # Construir mapa fila_admin -> fila_clasificado
    mapa = {}
    filas_elim = list(range(113, 129)) + list(range(145, 153)) + \
                 list(range(161, 165)) + list(range(169, 171))

    for r in filas_elim:
        fecha = ws.cell(row=r, column=COL_FECHA).value
        hora = ws.cell(row=r, column=COL_HORA).value
        if not fecha or not hora:
            continue
        fecha_str = fecha.strftime('%Y-%m-%d')
        hora_et = f'{(hora.hour + 1) % 24:02d}:{hora.strftime("%M")}'
        num_oficial = mapa_fecha_hora_a_num.get((fecha_str, hora_et))
        if num_oficial and num_oficial in clasif_por_num:
            mapa[r] = clasif_por_num[num_oficial]

    # Mapa especial para 3ro y Final (W101, W102, RU101, RU102, Terecro, Campeon)
    mapa_especial = {}
    for r in range(171, 179):
        label = ws.cell(row=r, column=COL_EQUIPO1).value
        if label:
            mapa_especial[str(label).strip()] = r

    return mapa, mapa_especial


def extraer_eliminatoria(ws, col_inicio, mapa_clasif, mapa_especial):
    """
    Lee los partidos de eliminatoria sumando correctamente los puntos del
    clasificado correspondiente a cada partido.
    """
    BLOQUES = [
        (range(113, 129), "Dieciseisavos"),
        (range(145, 153), "Octavos"),
        (range(161, 165), "Cuartos"),
        (range(169, 171), "Semifinales"),
        (range(175, 176), "3er y 4to puesto"),
        (range(177, 178), "Final"),
    ]

    MAPA_ESPECIAL_FASE = {
        "3er y 4to puesto": ("RU101", "RU102"),
        "Final": ("W101", "W102"),
    }

    rondas = []

    for bloque, nombre_fase in BLOQUES:
        for rr in bloque:
            num_partido = ws.cell(row=rr, column=COL_NUM_PARTIDO).value
            if num_partido is None:
                continue

            eq1 = ws.cell(row=rr, column=COL_EQUIPO1).value
            eq2 = ws.cell(row=rr, column=COL_EQUIPO2).value
            real_local  = ws.cell(row=rr, column=COL_REAL_LOCAL).value
            real_visit  = ws.cell(row=rr, column=COL_REAL_VISITANTE).value
            pred_local  = ws.cell(row=rr, column=col_inicio).value
            pred_visit  = ws.cell(row=rr, column=col_inicio + 1).value
            pred_ganador = ws.cell(row=rr, column=col_inicio + 2).value
            pts_goles   = ws.cell(row=rr, column=col_inicio + 3).value or 0

            jugado = real_local is not None and real_visit is not None

            # Ganador real del clasificado: columna 10 de la fila del PARTIDO
            real_ganador = ws.cell(row=rr, column=COL_REAL_GANADOR).value

            # Puntos y clasificado real desde las filas correspondientes
            pts_clasif = 0
            real_clasificado = None
            if nombre_fase in MAPA_ESPECIAL_FASE:
                # 3er puesto: RU101 + RU102 | Final: W101 + W102
                llave1, llave2 = MAPA_ESPECIAL_FASE[nombre_fase]
                fila_c1 = mapa_especial.get(llave1)
                fila_c2 = mapa_especial.get(llave2)
                # Sumar puntos de ambas filas de clasificado
                pts_c1 = ws.cell(row=fila_c1, column=col_inicio + 3).value or 0 if fila_c1 else 0
                pts_c2 = ws.cell(row=fila_c2, column=col_inicio + 3).value or 0 if fila_c2 else 0
                pts_clasif = pts_c1 + pts_c2
                # Clasificado real: viene de la fila 'Terecro' o 'Campeon' (col8),
                # NO de RU101/W101 que son los participantes del partido, no el ganador
                llave_resultado = 'Terecro' if nombre_fase == '3er y 4to puesto' else 'Campeon'
                fila_resultado = mapa_especial.get(llave_resultado)
                real_clasificado = ws.cell(row=fila_resultado, column=8).value if fila_resultado else None
            else:
                fila_c = mapa_clasif.get(rr)
                if fila_c:
                    pts_clasif = ws.cell(row=fila_c, column=col_inicio + 3).value or 0
                    real_clasificado = ws.cell(row=fila_c, column=8).value

            pts_total = pts_goles + pts_clasif

            eq2_invalido = (eq2 is None) or (isinstance(eq2, str) and eq2.strip().upper() == "#N/A")
            if eq2_invalido and eq1 in RIVAL_MEJOR_TERCERO:
                eq2_mostrar = RIVAL_MEJOR_TERCERO[eq1]
            elif eq2_invalido:
                eq2_mostrar = "Por definir"
            else:
                eq2_mostrar = eq2

            rondas.append({
                "ronda": nombre_fase,
                "equipo_local": eq1 if eq1 else "Por definir",
                "equipo_visitante": eq2_mostrar,
                "jugado": jugado,
                "resultado_real": {"local": real_local, "visitante": real_visit} if jugado else None,
                "clasificado_real": real_clasificado,
                "prediccion": {
                    "local": pred_local,
                    "visitante": pred_visit,
                    "ganador": pred_ganador,
                },
                "puntos": pts_total if jugado else None,
                "puntos_goles": pts_goles if jugado else None,
                "puntos_clasificado": pts_clasif if jugado else None,
            })

    return rondas


def extraer_premios(ws, col_inicio):
    premios = []
    r_start = None
    for r in range(FILA_FIN_GRUPOS, ws.max_row + 1):
        if ws.cell(row=r, column=COL_GRUPO).value == "Extra":
            r_start = r
            break
    if r_start is None:
        return premios

    r = r_start
    while r <= ws.max_row:
        label = ws.cell(row=r, column=COL_EQUIPO1).value
        if not label:
            r += 1
            if r - r_start > 12:
                break
            continue
        label_norm = PREMIOS_LABELS.get(label, label)
        prediccion = ws.cell(row=r, column=col_inicio + 2).value
        puntos = ws.cell(row=r, column=col_inicio + 3).value
        real = ws.cell(row=r, column=8).value  # col8 = valor real del premio
        premios.append({
            "premio": label_norm,
            "prediccion": prediccion,
            "real": real,
            "puntos": puntos,
        })
        r += 1
        if len(premios) >= len(PREMIOS_LABELS):
            break
    return premios


def exportar(path_admin, dir_salida, path_admin_json=None):
    wb = openpyxl.load_workbook(path_admin, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"No se encontro la hoja '{SHEET}' en {path_admin}")
    ws = wb[SHEET]

    os.makedirs(dir_salida, exist_ok=True)

    # Cargar JSON de partidos oficiales FIFA para construir mapa de clasificados
    if path_admin_json is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        path_admin_json = os.path.join(script_dir, "partidos_finales.json")

    if not os.path.exists(path_admin_json):
        raise FileNotFoundError(
            f"No se encontro {path_admin_json}. "
            "Asegurate de que partidos_finales.json este en la misma carpeta que este script."
        )

    participantes = detectar_participantes(ws)
    if not participantes:
        raise ValueError("No se encontraron bloques de participantes en la hoja Puntos")

    nombres_usados = {}
    indice = []

    with open(path_admin_json, encoding="utf-8") as f:
        partidos_pdf = json.load(f)

    mapa_clasif, mapa_especial = construir_mapa_clasificados(ws, partidos_pdf)

    for p in participantes:
        nombre = p["nombre"]
        col_inicio = p["col_inicio"]

        partidos_grupo = extraer_partidos_grupo(ws, col_inicio)
        eliminatoria = extraer_eliminatoria(ws, col_inicio, mapa_clasif, mapa_especial)
        premios = extraer_premios(ws, col_inicio)

        slug = slugify(nombre)
        if slug in nombres_usados:
            nombres_usados[slug] += 1
            slug = f"{slug}-{nombres_usados[slug]}"
        else:
            nombres_usados[slug] = 0

        data = {
            "nombre": nombre,
            "slug": slug,
            "generado": datetime.now().isoformat(timespec="seconds"),
            "fase_grupos": partidos_grupo,
            "eliminatoria": eliminatoria,
            "premios_fifa": premios,
        }

        path_out = os.path.join(dir_salida, f"{slug}.json")
        with open(path_out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

        indice.append({"nombre": nombre, "slug": slug})

    path_indice = os.path.join(dir_salida, "_indice.json")
    with open(path_indice, "w", encoding="utf-8") as f:
        json.dump({"generado": datetime.now().isoformat(timespec="seconds"),
                    "participantes": indice}, f, ensure_ascii=False, separators=(",", ":"))

    return indice


def main():
    if len(sys.argv) < 2:
        print("Uso: python exportar_predicciones.py <admin.xlsx> [dir_salida] [partidos_finales.json]")
        sys.exit(1)

    path_admin = sys.argv[1]
    dir_salida = sys.argv[2] if len(sys.argv) > 2 else "predicciones"
    path_json  = sys.argv[3] if len(sys.argv) > 3 else None

    indice = exportar(path_admin, dir_salida, path_json)

    print(f"Exportado: {len(indice)} archivos JSON en {dir_salida}/")
    print(f"Indice generado en {dir_salida}/_indice.json")


if __name__ == "__main__":
    main()
