"""
exportar_ranking.py

Lee la hoja 'Posiciones' del archivo Admin de la Polla Mundialista y genera
un JSON estatico con el historico de puntos por jornada (J0..Jn) para cada
participante. Este JSON alimenta la grafica animada (bar chart race) en la
web publicada en GitHub Pages.

NO recalcula nada: solo lee valores ya calculados por las formulas de Excel
en el Admin. El Admin sigue siendo la unica fuente de verdad.

Uso:
    python exportar_ranking.py "Admin_POLLA_FIFA_WORLD_CUP_2026_r00.xlsx" "salida/ranking_historico.json"

Si no se indica ruta de salida, escribe en ./ranking_historico.json
"""

import sys
import json
import re
from datetime import datetime
import openpyxl


SHEET = "Posiciones"
FILA_HEADERS = 3
FILA_FECHAS = 2
PRIMERA_FILA_PARTICIPANTE = 4
COL_NOMBRE = 15  # columna O: "Nombre" (bloque derecho de la hoja)


def detectar_columnas_jornada(ws):
    """
    Recorre la fila de headers buscando columnas con patron J<numero>
    (J0, J1, J2, ...) y devuelve lista de tuplas (col, jornada_idx, fecha)
    ordenadas por jornada_idx ascendente. No asume un numero fijo de
    jornadas: lee tantas como existan en el archivo.
    """
    patron = re.compile(r"^J(\d+)$")
    columnas = []
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=FILA_HEADERS, column=c).value
        if not header or not isinstance(header, str):
            continue
        m = patron.match(header.strip())
        if not m:
            continue
        jornada_idx = int(m.group(1))
        fecha = ws.cell(row=FILA_FECHAS, column=c).value
        columnas.append((c, jornada_idx, fecha))
    columnas.sort(key=lambda x: x[1])
    return columnas


def detectar_ultima_jornada_con_datos(ws, columnas, primera_fila, ultima_fila):
    """
    Devuelve el jornada_idx mas alto que tenga al menos un valor numerico
    no nulo en cualquier participante. Las jornadas futuras (sin jugar)
    tienen celdas vacias y no deben incluirse en el JSON final, para que
    la animacion no muestre jornadas "fantasma" con puntos en cero.
    """
    ultima = 0
    for col, jornada_idx, _fecha in columnas:
        tiene_datos = False
        for r in range(primera_fila, ultima_fila + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                tiene_datos = True
                break
        if tiene_datos:
            ultima = max(ultima, jornada_idx)
    return ultima


def slugify(nombre):
    s = nombre.strip().lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
           .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def exportar(path_admin, path_salida):
    wb = openpyxl.load_workbook(path_admin, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"No se encontro la hoja '{SHEET}' en {path_admin}")
    ws = wb[SHEET]

    columnas = detectar_columnas_jornada(ws)
    if not columnas:
        raise ValueError("No se encontraron columnas con patron J<numero> en la hoja Posiciones")

    ultima_fila = PRIMERA_FILA_PARTICIPANTE
    while ws.cell(row=ultima_fila + 1, column=COL_NOMBRE).value:
        ultima_fila += 1

    ultima_jornada = detectar_ultima_jornada_con_datos(
        ws, columnas, PRIMERA_FILA_PARTICIPANTE, ultima_fila
    )
    columnas_validas = [c for c in columnas if c[1] <= ultima_jornada]

    participantes = []
    nombres_usados = {}
    for r in range(PRIMERA_FILA_PARTICIPANTE, ultima_fila + 1):
        nombre = ws.cell(row=r, column=COL_NOMBRE).value
        if not nombre:
            continue
        nombre = str(nombre).strip()

        puntos = []
        for col, _jornada_idx, _fecha in columnas_validas:
            v = ws.cell(row=r, column=col).value
            puntos.append(v if isinstance(v, (int, float)) else 0)

        slug = slugify(nombre)
        if slug in nombres_usados:
            nombres_usados[slug] += 1
            slug = f"{slug}-{nombres_usados[slug]}"
        else:
            nombres_usados[slug] = 0

        participantes.append({
            "nombre": nombre,
            "slug": slug,
            "puntos": puntos,
        })

    fechas = []
    for _col, jornada_idx, fecha in columnas_validas:
        if isinstance(fecha, datetime):
            fechas.append(fecha.strftime("%Y-%m-%d"))
        else:
            fechas.append(None)

    salida = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "ultima_jornada": ultima_jornada,
        "jornadas": [c[1] for c in columnas_validas],
        "fechas": fechas,
        "participantes": participantes,
    }

    with open(path_salida, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, separators=(",", ":"))

    return salida


def main():
    if len(sys.argv) < 2:
        print("Uso: python exportar_ranking.py <admin.xlsx> [salida.json]")
        sys.exit(1)

    path_admin = sys.argv[1]
    path_salida = sys.argv[2] if len(sys.argv) > 2 else "ranking_historico.json"

    salida = exportar(path_admin, path_salida)

    print(f"Exportado: {path_salida}")
    print(f"Participantes: {len(salida['participantes'])}")
    print(f"Jornadas incluidas: J0 a J{salida['ultima_jornada']}")


if __name__ == "__main__":
    main()
