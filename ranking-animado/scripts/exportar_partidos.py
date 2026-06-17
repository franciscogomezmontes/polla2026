"""
exportar_partidos.py

Lee la hoja 'Puntos' del Admin de la Polla Mundialista y genera un archivo
JSON por partido (104 en total: 72 de fase de grupos + 32 de eliminatoria)
con: equipos, fecha/hora, ciudad sede, resultado real (si ya se jugo), y
la prediccion + puntos de cada uno de los participantes, ordenados segun
su posicion en el ranking de la jornada correspondiente a la fecha del
partido.

NO recalcula puntos: los toma directamente de la columna ya calculada por
formulas en el Admin (col_base+3). El Admin sigue siendo la unica fuente
de verdad para reglas y puntuacion.

La ciudad sede no existe en el Admin: se cruza usando el diccionario
codigo_a_nombre_admin.json (codigo FIFA de 3 letras -> nombre exacto del
Admin) contra el calendario oficial ya extraido en partidos_finales.json.
Ambos archivos deben estar en la misma carpeta que este script, o se
indica su ruta con --partidos-pdf / --codigo-nombre.

Requiere que ranking_historico.json ya este generado (correr primero
exportar_ranking.py), para poder ordenar a los participantes por su
ranking real en la jornada de cada partido.

Uso:
    python exportar_partidos.py "Admin_POLLA_FIFA_WORLD_CUP_2026_r00.xlsx" "salida/partidos/" "ranking_historico.json"
"""

import sys
import os
import json
import re
from datetime import datetime, time as dtime
import openpyxl


SHEET = "Puntos"
FILA_NOMBRE_PART = 3
FILA_TIPO_COL = 4
PRIMER_COL_PARTICIPANTE = 11
COLS_POR_BLOQUE = 4

COL_NUM_PARTIDO = 1
COL_GRUPO = 2
COL_FECHA = 3
COL_HORA = 4
COL_EQUIPO1 = 5
COL_VS = 6
COL_EQUIPO2 = 7
COL_REAL_LOCAL = 8
COL_REAL_VISITANTE = 9
COL_REAL_GANADOR = 10

FILA_INICIO_GRUPOS = 5
FILA_FIN_GRUPOS = 112
FILA_INICIO_ELIM = 113
FILA_FIN_ELIM = 180

FASES_ELIMINATORIA = {
    "DIECISEISAVOS": "Dieciseisavos",
    "OCTAVOS": "Octavos",
    "CUARTOS": "Cuartos",
    "SEMIS": "Semifinales",
    "3ro": "3er y 4to puesto",
    "1ro": "Final",
}


RIVAL_MEJOR_TERCERO = {
    "1E": "3ABCDF",
    "1I": "3CDFGH",
    "1A": "3CEFHI",
    "1L": "3EHIJK",
    "1D": "3BEFIJ",
    "1G": "3AEHIJ",
    "1B": "3EFGIJ",
    "1K": "3DEIJL",
}


def slugify(nombre):
    s = nombre.strip().lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
           .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def fmt_fecha(fecha):
    if isinstance(fecha, datetime):
        return fecha.strftime("%Y-%m-%d")
    return None


def fmt_hora(hora):
    if isinstance(hora, dtime):
        return hora.strftime("%H:%M")
    return None


def detectar_participantes(ws):
    participantes = []
    c = PRIMER_COL_PARTICIPANTE
    while c <= ws.max_column:
        nombre = ws.cell(row=FILA_NOMBRE_PART, column=c).value
        if not nombre:
            break
        participantes.append({"nombre": str(nombre).strip(), "col_inicio": c})
        c += COLS_POR_BLOQUE
    return participantes


def cargar_ciudades(ruta_partidos_pdf, ruta_codigo_nombre):
    """
    Devuelve tres indices sobre los partidos del calendario oficial (PDF):
    - por_equipos: frozenset({nombre_admin_eq1, nombre_admin_eq2}) -> partido
      (para partidos de fase de grupos, donde los equipos ya son nombres reales)
    - por_codigo_posicion: frozenset({"1A","2B"}) -> partido
      (para Dieciseisavos, donde los equipos son codigos de posicion de grupo)
    - por_hora: "HH:MM" -> lista de partidos con esa hora
      (para Octavos en adelante, donde no hay equipos fijos: se cruza por
      fecha+hora con tolerancia de 1h, ya validado contra el Admin)
    """
    with open(ruta_partidos_pdf, encoding="utf-8") as f:
        partidos_pdf = json.load(f)
    with open(ruta_codigo_nombre, encoding="utf-8") as f:
        cod2nombre = json.load(f)

    por_equipos = {}
    por_codigo_posicion = {}
    por_hora = {}
    for p in partidos_pdf:
        eq1, eq2 = p.get("equipo1"), p.get("equipo2")
        if eq1 and eq2:
            n1, n2 = cod2nombre.get(eq1), cod2nombre.get(eq2)
            if n1 and n2:
                key = frozenset({n1.strip(), n2.strip()})
                por_equipos[key] = p
            else:
                key2 = frozenset({eq1, eq2})
                por_codigo_posicion[key2] = p
        if p.get("hora"):
            por_hora.setdefault(p["hora"], []).append(p)

    return por_equipos, por_codigo_posicion, por_hora


def hora_menos_una(hora_str):
    """'13:00' -> '12:00'. Usado para cruzar hora Colombia (Admin) con
    hora Eastern Time en horario de verano (PDF), offset confirmado de 1h."""
    if not hora_str:
        return None
    h, m = hora_str.split(":")
    h = (int(h) - 1) % 24
    return f"{h:02d}:{m}"


def buscar_partido_pdf_por_hora(hora_admin, por_hora, usados):
    """
    Busca en por_hora el partido cuya hora ET sea (hora_admin + 1h),
    excluyendo los numero_oficial ya usados (para evitar que dos
    partidos del Admin el mismo dia a la misma hora-base colisionen).
    """
    hora_et_esperada = None
    if hora_admin:
        h, m = hora_admin.split(":")
        hora_et_esperada = f"{(int(h)+1)%24:02d}:{m}"
    candidatos = por_hora.get(hora_et_esperada, [])
    for c in candidatos:
        if c["numero"] not in usados:
            return c
    return None


def buscar_ciudad(eq1_admin, eq2_admin, por_equipos, por_codigo_posicion, cod2nombre_inv):
    """
    Intenta encontrar la info de ciudad/numero oficial para un partido,
    primero por nombre real de equipo, luego por codigo de posicion
    (1A, 2B, 3CEFHI, etc, que es como aparecen los partidos de
    eliminatoria antes de que se conozcan los equipos reales).
    """
    if eq1_admin and eq2_admin:
        key = frozenset({str(eq1_admin).strip(), str(eq2_admin).strip()})
        if key in por_equipos:
            return por_equipos[key]
        key2 = frozenset({str(eq1_admin).strip(), str(eq2_admin).strip()})
        if key2 in por_codigo_posicion:
            return por_codigo_posicion[key2]
    return None


def cargar_ranking_historico(ruta):
    with open(ruta, encoding="utf-8") as f:
        data = json.load(f)
    return data


def jornada_para_fecha(fecha_str, ranking_data):
    """
    Devuelve el indice de jornada (0,1,2...) cuya fecha es la ultima
    fecha <= fecha_str. Si el partido es anterior a J0, devuelve 0.
    Si no hay fechas registradas, devuelve None.
    """
    fechas = ranking_data.get("fechas", [])
    if not fechas or not fecha_str:
        return None
    idx_valido = None
    for i, f in enumerate(fechas):
        if f and f <= fecha_str:
            idx_valido = i
    if idx_valido is None:
        return 0
    return idx_valido


def construir_ranking_para_jornada(ranking_data, jornada_idx):
    """
    Devuelve lista de {nombre, slug, puntos, rank} ordenada por ranking
    real (con empates) para la jornada indicada.
    """
    if jornada_idx is None:
        jornada_idx = len(ranking_data.get("jornadas", [])) - 1

    items = []
    for p in ranking_data["participantes"]:
        pts = p["puntos"][jornada_idx] if jornada_idx < len(p["puntos"]) else 0
        items.append({"nombre": p["nombre"], "slug": p["slug"], "puntos": pts})

    items.sort(key=lambda x: -x["puntos"])
    ranked = []
    last_pts, last_rank = None, 0
    for i, x in enumerate(items):
        if x["puntos"] == last_pts:
            rank = last_rank
        else:
            rank = i + 1
            last_rank = rank
            last_pts = x["puntos"]
        ranked.append({**x, "rank": rank})
    return ranked


def extraer_filas_predecesoras(ws):
    """
    Recorre toda la hoja y construye un diccionario
    {numero_oficial_fifa: fila} para cada fila tipo "Clasificado MXX",
    "Clasificado WXX", "WXX" o "RUXX" (texto en COL_EQUIPO1), que
    representan la prediccion de cada participante sobre que equipo
    avanza de un partido anterior. La clave es el numero oficial FIFA
    como entero (ej. "Clasificado W93" -> 93, "RU101" -> guarda aparte
    bajo clave "RU101" porque RU no es directamente un numero de partido
    sino "runner-up de la semifinal 101/102", caso especial).
    """
    predecesoras = {}
    for r in range(FILA_INICIO_ELIM, FILA_FIN_ELIM):
        val = ws.cell(row=r, column=COL_EQUIPO1).value
        if not val or not isinstance(val, str):
            continue
        m = re.match(r"^Clasificado [MW](\d+)$", val.strip())
        if m:
            predecesoras[f"W{m.group(1)}"] = r
            continue
        m2 = re.match(r"^(W|RU)(\d+)$", val.strip())
        if m2:
            predecesoras[val.strip()] = r
    return predecesoras


def extraer_partidos_admin(ws):
    """
    Recorre la hoja Puntos completa (grupos + eliminatoria) y devuelve
    una lista de partidos con su info base, SIN predicciones todavia.
    Para partidos de Octavos en adelante (donde equipo1/equipo2 estan
    vacios en la fila del partido real), guarda en su lugar las filas
    predecesoras de donde se debe leer la prediccion de cada equipo
    por participante (ver "filas_equipo1"/"filas_equipo2").
    """
    partidos = []
    grupo_actual = None

    r = FILA_INICIO_GRUPOS
    while r <= FILA_FIN_GRUPOS:
        grupo_cell = ws.cell(row=r, column=COL_GRUPO).value
        if grupo_cell:
            grupo_actual = grupo_cell

        num_partido = ws.cell(row=r, column=COL_NUM_PARTIDO).value
        eq1 = ws.cell(row=r, column=COL_EQUIPO1).value

        if num_partido is not None and eq1 is not None and not re.match(r"^\d[A-L]$", str(eq1)):
            eq2 = ws.cell(row=r, column=COL_EQUIPO2).value
            fecha = ws.cell(row=r, column=COL_FECHA).value
            hora = ws.cell(row=r, column=COL_HORA).value
            real_local = ws.cell(row=r, column=COL_REAL_LOCAL).value
            real_visit = ws.cell(row=r, column=COL_REAL_VISITANTE).value
            jugado = real_local is not None and real_visit is not None

            partidos.append({
                "fase": "Fase de grupos",
                "grupo": grupo_actual,
                "fila": r,
                "equipo1": eq1,
                "equipo2": eq2,
                "fila_equipo1": None,
                "fila_equipo2": None,
                "fecha": fmt_fecha(fecha),
                "hora": fmt_hora(hora),
                "jugado": jugado,
                "resultado_real": {"local": real_local, "visitante": real_visit} if jugado else None,
            })
        r += 1

    predecesoras = extraer_filas_predecesoras(ws)

    fase_actual = None
    r = FILA_INICIO_ELIM
    while r <= FILA_FIN_ELIM:
        fase_cell = ws.cell(row=r, column=COL_GRUPO).value
        if fase_cell in FASES_ELIMINATORIA:
            fase_actual = FASES_ELIMINATORIA[fase_cell]

        num_partido = ws.cell(row=r, column=COL_NUM_PARTIDO).value
        eq1 = ws.cell(row=r, column=COL_EQUIPO1).value

        if num_partido is not None and fase_actual and (eq1 is not None or fase_actual != "Dieciseisavos"):
            if fase_actual == "Dieciseisavos":
                eq2 = ws.cell(row=r, column=COL_EQUIPO2).value
                if eq2 == "#N/A" and eq1 in RIVAL_MEJOR_TERCERO:
                    eq2 = RIVAL_MEJOR_TERCERO[eq1]
                fila_eq1, fila_eq2 = None, None
            else:
                eq2 = None
                fila_eq1, fila_eq2 = None, None

            fecha = ws.cell(row=r, column=COL_FECHA).value
            hora = ws.cell(row=r, column=COL_HORA).value
            real_local = ws.cell(row=r, column=COL_REAL_LOCAL).value
            real_visit = ws.cell(row=r, column=COL_REAL_VISITANTE).value
            jugado = real_local is not None and real_visit is not None

            partidos.append({
                "fase": fase_actual,
                "grupo": None,
                "fila": r,
                "equipo1": eq1,
                "equipo2": eq2,
                "fila_equipo1": fila_eq1,
                "fila_equipo2": fila_eq2,
                "fecha": fmt_fecha(fecha),
                "hora": fmt_hora(hora),
                "jugado": jugado,
                "resultado_real": {"local": real_local, "visitante": real_visit} if jugado else None,
            })
        r += 1

    return partidos, predecesoras


def main():
    if len(sys.argv) < 4:
        print("Uso: python exportar_partidos.py <admin.xlsx> <carpeta_salida> <ranking_historico.json> "
              "[--partidos-pdf partidos_finales.json] [--codigo-nombre codigo_a_nombre_admin.json]")
        sys.exit(1)

    ruta_admin = sys.argv[1]
    carpeta_salida = sys.argv[2]
    ruta_ranking = sys.argv[3]

    script_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_partidos_pdf = os.path.join(script_dir, "partidos_finales.json")
    ruta_codigo_nombre = os.path.join(script_dir, "codigo_a_nombre_admin.json")

    for i, arg in enumerate(sys.argv):
        if arg == "--partidos-pdf" and i + 1 < len(sys.argv):
            ruta_partidos_pdf = sys.argv[i + 1]
        if arg == "--codigo-nombre" and i + 1 < len(sys.argv):
            ruta_codigo_nombre = sys.argv[i + 1]

    if not os.path.exists(ruta_admin):
        print(f"No se encontro: {ruta_admin}")
        sys.exit(1)
    if not os.path.exists(ruta_partidos_pdf):
        print(f"No se encontro: {ruta_partidos_pdf}")
        sys.exit(1)
    if not os.path.exists(ruta_codigo_nombre):
        print(f"No se encontro: {ruta_codigo_nombre}")
        sys.exit(1)
    if not os.path.exists(ruta_ranking):
        print(f"No se encontro: {ruta_ranking}")
        print("Corre primero exportar_ranking.py para generar ranking_historico.json")
        sys.exit(1)

    os.makedirs(carpeta_salida, exist_ok=True)

    print(f"Leyendo {ruta_admin}...")
    wb = openpyxl.load_workbook(ruta_admin, data_only=True)
    ws = wb[SHEET]

    participantes = detectar_participantes(ws)
    print(f"Participantes detectados: {len(participantes)}")

    print("Cargando ciudades desde calendario oficial...")
    por_equipos, por_codigo_posicion, por_hora = cargar_ciudades(ruta_partidos_pdf, ruta_codigo_nombre)

    print("Cargando ranking historico...")
    ranking_data = cargar_ranking_historico(ruta_ranking)

    print("Extrayendo partidos de la hoja Puntos...")
    partidos_base, predecesoras = extraer_partidos_admin(ws)
    print(f"Partidos encontrados: {len(partidos_base)}")

    indice = []
    sin_ciudad = []
    numeros_oficiales_usados = set()

    # Fases donde equipo1/equipo2 dependen de la prediccion de cada
    # participante (vienen de filas "Clasificado WXX" / "WXX" / "RUXX")
    FASES_BRACKET_PERSONALIZADO = {"Octavos", "Cuartos", "Semifinales", "3er y 4to puesto", "Final"}

    for n, partido in enumerate(partidos_base, start=1):
        es_bracket_personalizado = partido["fase"] in FASES_BRACKET_PERSONALIZADO

        if not es_bracket_personalizado:
            info_ciudad = buscar_ciudad(partido["equipo1"], partido["equipo2"], por_equipos, por_codigo_posicion, None)
            ciudad = info_ciudad["ciudad"] if info_ciudad else None
            numero_oficial = info_ciudad["numero"] if info_ciudad else None
        else:
            info_ciudad = buscar_partido_pdf_por_hora(partido["hora"], por_hora, numeros_oficiales_usados)
            ciudad = info_ciudad["ciudad"] if info_ciudad else None
            numero_oficial = info_ciudad["numero"] if info_ciudad else None

        if numero_oficial is not None:
            numeros_oficiales_usados.add(numero_oficial)
        if not ciudad:
            sin_ciudad.append((n, partido["equipo1"], partido["equipo2"]))

        jornada_idx = jornada_para_fecha(partido["fecha"], ranking_data)
        ranking_jornada = construir_ranking_para_jornada(ranking_data, jornada_idx)

        fila_pred_eq1 = None
        fila_pred_eq2 = None
        if partido["fase"] == "Final":
            fila_pred_eq1 = predecesoras.get("W101")
            fila_pred_eq2 = predecesoras.get("W102")
        elif partido["fase"] == "3er y 4to puesto":
            fila_pred_eq1 = predecesoras.get("RU101")
            fila_pred_eq2 = predecesoras.get("RU102")
        elif es_bracket_personalizado and info_ciudad:
            cod_eq1, cod_eq2 = info_ciudad.get("equipo1"), info_ciudad.get("equipo2")
            fila_pred_eq1 = predecesoras.get(cod_eq1) if cod_eq1 else None
            fila_pred_eq2 = predecesoras.get(cod_eq2) if cod_eq2 else None

        predicciones = []
        for p in participantes:
            col = p["col_inicio"]
            fila = partido["fila"]
            pred_local = ws.cell(row=fila, column=col).value
            pred_visit = ws.cell(row=fila, column=col + 1).value
            pred_ganador = ws.cell(row=fila, column=col + 2).value
            pred_puntos = ws.cell(row=fila, column=col + 3).value

            equipo1_part = partido["equipo1"]
            equipo2_part = partido["equipo2"]
            if es_bracket_personalizado:
                equipo1_part = ws.cell(row=fila_pred_eq1, column=col + 2).value if fila_pred_eq1 else None
                equipo2_part = ws.cell(row=fila_pred_eq2, column=col + 2).value if fila_pred_eq2 else None

            slug = slugify(p["nombre"])
            rank_info = next((r for r in ranking_jornada if r["slug"] == slug), None)

            predicciones.append({
                "nombre": p["nombre"],
                "slug": slug,
                "rank": rank_info["rank"] if rank_info else None,
                "equipo1": equipo1_part,
                "equipo2": equipo2_part,
                "prediccion": {
                    "local": pred_local,
                    "visitante": pred_visit,
                    "ganador": pred_ganador,
                },
                "puntos": pred_puntos if partido["jugado"] else None,
            })

        predicciones.sort(key=lambda x: (x["rank"] if x["rank"] is not None else 9999))

        partido_out = {
            "numero": n,
            "numero_oficial_fifa": numero_oficial,
            "fase": partido["fase"],
            "grupo": partido["grupo"],
            "equipos_fijos": not es_bracket_personalizado,
            "equipo1": partido["equipo1"],
            "equipo2": partido["equipo2"],
            "fecha": partido["fecha"],
            "hora": partido["hora"],
            "ciudad": ciudad,
            "jugado": partido["jugado"],
            "resultado_real": partido["resultado_real"],
            "predicciones": predicciones,
        }

        with open(os.path.join(carpeta_salida, f"partido_{n}.json"), "w", encoding="utf-8") as f:
            json.dump(partido_out, f, ensure_ascii=False)

        indice.append({
            "numero": n,
            "fase": partido["fase"],
            "grupo": partido["grupo"],
            "equipos_fijos": not es_bracket_personalizado,
            "equipo1": partido["equipo1"],
            "equipo2": partido["equipo2"],
            "fecha": partido["fecha"],
            "hora": partido["hora"],
            "ciudad": ciudad,
            "jugado": partido["jugado"],
        })

    with open(os.path.join(carpeta_salida, "_indice.json"), "w", encoding="utf-8") as f:
        json.dump({"generado": datetime.now().isoformat(), "partidos": indice}, f, ensure_ascii=False)

    print(f"\nExportados {len(partidos_base)} archivos partido_N.json en {carpeta_salida}")
    print(f"Indice generado en {os.path.join(carpeta_salida, '_indice.json')}")
    if sin_ciudad:
        print(f"\nAviso: {len(sin_ciudad)} partido(s) sin ciudad encontrada:")
        for n, e1, e2 in sin_ciudad:
            print(f"  - Partido {n}: {e1} vs {e2}")


if __name__ == "__main__":
    main()
